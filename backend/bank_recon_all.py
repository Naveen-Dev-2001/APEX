"""
Sage Intacct Bank Reconciliation — All Uncleared Transactions Fetcher
===============================================================

A transaction can land on the Bank Reconciliation screen through FIVE
different Intacct workflows — four are credits/inflows to the bank
account, one (APPYMT) is a debit/outflow:

  1. Cash Management > Fund Transfer            -> object FUNDSTRANSFER
  2. Cash Management > Other Receipts           -> object OTHERRECEIPTS
  3. Accounts Receivable > Deposits             -> object DEPOSIT
  4. Accounts Receivable > Receive Payment
       (payment method = Check)                 -> object ARPYMT
  5. Accounts Payable > Pay Bills
       (checks/debits OUT to vendors)           -> object APPYMT

APPYMT is the money-OUT counterpart to ARPYMT: same shape (FINANCIALENTITY,
CLEARED, BANKTRXAMOUNT, DOCNUMBER, PAYMENTMETHOD) but the party is a
VENDORNAME instead of a CUSTOMERNAME, and on the bank statement it will
show up as a debit rather than a credit. CONFIRMED against a real sandbox
lookup + query (see lookup_APPYMT.xml / debug_page1_appymt.xml) — all
fields below are real, not guessed. Note: BANKTRXAMOUNT on APPYMT is
POSITIVE even though it's an outflow (e.g. 14626.52 for a vendor check) —
Intacct does not sign-flip it, so downstream matching against the bank
statement's debit/credit columns must key off the "direction" convenience
flag this script adds, not the sign of the amount.

This script:
  1. Authenticates and gets a session id.
  2. Runs a <lookup> against each object FIRST and saves the raw response
     to lookup_<OBJECT>.xml — use this to confirm real field names before
     trusting the filtered query results below. This matters most for
     FUNDSTRANSFER, which likely does NOT have a single FINANCIALENTITY
     field the way ARPYMT/DEPOSIT/OTHERRECEIPTS do.
  3. Runs readByQuery + readMore pagination against each object, filtered
     to a given financial entity and CLEARED = 'F' (uncleared only).
  4. Normalizes + saves results to JSON, and writes an .xlsx with columns
     source_object, doc_number, description, total, party_name for the
     uncleared (CLEARED='F') rows.

If this legacy XML gateway approach doesn't surface what you see on the
bank rec screen, the next step is the newer REST API with OAuth2 —
flag that separately once we've confirmed (or ruled out) the XML path.
"""

import httpx
import xmltodict
import asyncio
import json
import os
from xml.sax.saxutils import escape as xml_escape

SAGE_URL = "https://api.intacct.com/ia/xml/xmlgw.phtml"


def load_env_file(path: str = ".env") -> None:
    """
    Minimal .env loader (no external dependency). Reads KEY=VALUE lines,
    strips surrounding quotes/whitespace, and only sets a variable if it
    isn't already set in the real OS environment (so a real env var always
    wins over the file). Silently does nothing if the file doesn't exist.

    Note: the file must be named exactly ".env" and live in the same
    folder you run the script from (or pass a full path in).
    """
    if not os.path.exists(path):
        print(f"  ⚠ No .env file found at '{os.path.abspath(path)}' — "
              f"falling back to CHANGE_ME placeholders if OS env vars aren't set.")
        return
    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            os.environ.setdefault(key, value)
    print(f"  ✓ Loaded credentials from {os.path.abspath(path)}")


load_env_file()

# ── Credentials ──────────────────────────────────────────────────────────
# SECURITY NOTE: these were hardcoded in plaintext in an earlier version of
# this script. That password should be considered exposed — please rotate
# it in Intacct regardless of the .env fix below.
SENDER_ID       = os.environ.get("SAGE_SENDER_ID", "consolidatedanalytic")
SENDER_PASSWORD = os.environ.get("SAGE_SENDER_PASSWORD", "CHANGE_ME")
USER_ID         = os.environ.get("SAGE_USER_ID", "Apex")
LOCATION_ID     = os.environ.get("SAGE_LOCATION_ID", "")   # e.g. "303" — optional, only for multi-entity/location-scoped logins
COMPANY_ID      = os.environ.get("SAGE_COMPANY_ID", "consolidatedanalytics-sandbox")
USER_PASSWORD   = os.environ.get("SAGE_USER_PASSWORD", "CHANGE_ME")

if SENDER_PASSWORD == "CHANGE_ME" or USER_PASSWORD == "CHANGE_ME":
    print("  ✗ Credentials still on CHANGE_ME placeholder — check your .env "
          "file name/location and contents before running further.")

# Start with FFB_4449 only, per your instruction. Add FFB_8723 back in
# once FFB_4449 is confirmed working.
FINANCIAL_ENTITIES = ["FFB_4183"]

# object_name -> lowercase XML record tag xmltodict will produce
OBJECTS = {
    "FUNDSTRANSFER": "fundstransfer",
    "OTHERRECEIPTS": "otherreceipts",
    "DEPOSIT":       "deposit",
    "ARPYMT":        "arpymt",
    "APPYMT":        "appymt",
}

# object_name -> field to use for the diagnostic sample (confirmed via lookup)
ACCOUNT_FIELD = {
    "FUNDSTRANSFER": "TOACCOUNTID",   # also has FROMACCOUNTID and TOACCOUNTID
    "OTHERRECEIPTS": "BANKACCOUNTID",
    "DEPOSIT":       "FINANCIALENTITY",
    "ARPYMT":        "FINANCIALENTITY",
    "APPYMT":        "FINANCIALENTITY",   # confirmed via lookup_APPYMT.xml + debug_page1_appymt.xml
}


# ── Auth ──────────────────────────────────────────────────────────────────

async def get_session_id() -> str:
    # <locationid> is a separate element in the <login> block for
    # multi-entity/location-scoped Intacct companies — it is NOT combined
    # into <userid> with a pipe character. That "userid|location" pattern
    # you may have seen elsewhere doesn't apply to this XML gateway login
    # block, which is why it previously came back as a literal (invalid)
    # userid "Apex|303" with locationid still None in the failure response.
    location_xml = f"        <locationid>{LOCATION_ID}</locationid>\n" if LOCATION_ID else ""

    payload = f"""<?xml version="1.0" encoding="UTF-8"?>
<request>
  <control>
    <senderid>{SENDER_ID}</senderid>
    <password>{SENDER_PASSWORD}</password>
    <controlid>get-session</controlid>
    <uniqueid>false</uniqueid>
    <dtdversion>3.0</dtdversion>
  </control>
  <operation>
    <authentication>
      <login>
        <userid>{USER_ID}</userid>
        <companyid>{COMPANY_ID}</companyid>
        <password>{USER_PASSWORD}</password>
{location_xml}
      </login>
    </authentication>
    <content>
      <function controlid="session">
        <getAPISession/>
      </function>
    </content>
  </operation>
</request>"""

    print("  → Getting session token...")
    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.post(
            SAGE_URL, content=payload,
            headers={"Content-Type": "application/xml"}
        )

    with open("debug_get_session.xml", "w") as f:
        f.write(response.text)

    data = xmltodict.parse(response.text)
    resp = data.get("response", {})

    # ── Control-level failure (bad senderid/senderpassword, malformed
    #    request, etc.) — this response has NO "operation" key, which is
    #    what caused the earlier KeyError. Surface the real reason instead.
    control = resp.get("control", {})
    if control.get("status") == "failure" or "operation" not in resp:
        err = resp.get("errormessage", {}).get("error", {})
        if isinstance(err, list):
            err = err[0]
        detail = err.get("description2") or err.get("description") or err or "no error detail returned"
        raise Exception(
            f"Sage control-level failure while getting session: {detail}\n"
            f"HTTP status: {response.status_code}\n"
            f"Raw response saved to debug_get_session.xml — check it directly.\n"
            f"Most common causes: SENDER_ID/SENDER_PASSWORD wrong or unset "
            f"(currently using env var if set, else 'CHANGE_ME' placeholder), "
            f"or USER_ID/COMPANY_ID/USER_PASSWORD wrong."
        )

    op = resp["operation"]

    if op["authentication"]["status"] != "success":
        raise Exception(f"Auth failed: {op['authentication']}")

    result = op["result"]
    if result["status"] != "success":
        raise Exception(f"Session error: {result['errormessage']['error']['description2']}")

    session_id = result["data"]["api"]["sessionid"]
    print(f"  ✓ Session obtained: {session_id[:20]}...")
    return session_id


# ── Field discovery via <lookup> ───────────────────────────────────────────

async def run_lookup(session_id: str, object_name: str) -> list[str]:
    """
    Runs <lookup><object>NAME</object></lookup> to get the real field list
    for an object. Saves the raw XML so you can inspect it directly, and
    prints any field whose name contains "FINANC" as a hint for which
    field to filter on.
    """
    payload = f"""<?xml version="1.0" encoding="UTF-8"?>
<request>
  <control>
    <senderid>{SENDER_ID}</senderid>
    <password>{SENDER_PASSWORD}</password>
    <controlid>lookup-{object_name}</controlid>
    <uniqueid>false</uniqueid>
    <dtdversion>3.0</dtdversion>
  </control>
  <operation>
    <authentication>
      <sessionid>{session_id}</sessionid>
    </authentication>
    <content>
      <function controlid="lookup-{object_name}">
        <lookup>
          <object>{object_name}</object>
        </lookup>
      </function>
    </content>
  </operation>
</request>"""

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.post(
            SAGE_URL, content=payload,
            headers={"Content-Type": "application/xml"}
        )

    debug_path = f"lookup_{object_name}.xml"
    with open(debug_path, "w") as f:
        f.write(response.text)

    data = xmltodict.parse(response.text)
    result = data.get("response", {}).get("operation", {}).get("result", {})

    if result.get("status") == "failure":
        err = result.get("errormessage", {}).get("error", {})
        print(f"  ✗ [{object_name}] lookup failed: {err.get('description2', err)} "
              f"(see {debug_path})")
        return []

    try:
        fields = result["data"]["Type"]["Fields"]["Field"]
    except (KeyError, TypeError):
        print(f"  ⚠ [{object_name}] lookup returned unexpected shape — see {debug_path}")
        return []

    if isinstance(fields, dict):
        fields = [fields]

    field_names = [f.get("ID") for f in fields if isinstance(f, dict) and f.get("ID")]
    print(f"  ✓ [{object_name}] {len(field_names)} fields — saved to {debug_path}")

    fe_candidates = [f for f in field_names if "FINANC" in f.upper() or "BANK" in f.upper() or "ACCOUNT" in f.upper()]
    if fe_candidates:
        print(f"      possible account/financial-entity fields: {fe_candidates}")

    return field_names


# ── Fetch with readByQuery + readMore pagination ───────────────────────────

async def fetch_object_records(
    session_id: str,
    object_name: str,
    record_tag: str,
    query_str: str,
    page_size: int = 100,
) -> list[dict]:
    """Generic readByQuery + readMore pagination fetcher for any Intacct object."""

    all_records = []
    page_num = 1
    result_id = None
    total_count = None

    # Escape XML special characters (<, >, &) in the query string — Sage's
    # query language uses operators like "<>" or "!=" for not-equal, and
    # "<" / ">" for numeric comparisons, but those characters are also XML
    # syntax. Left unescaped, the gateway's XML parser sees a literal "<"
    # and thinks a new tag is starting, causing an "invalid element name"
    # schema error.
    safe_query_str = xml_escape(query_str)

    async with httpx.AsyncClient(timeout=60) as client:
        while True:
            print(f"  → [{object_name}] Fetching page {page_num}...")

            if result_id is None:
                function_xml = f"""
                    <function controlid="page-{page_num}">
                      <readByQuery>
                        <object>{object_name}</object>
                        <fields>*</fields>
                        <query>{safe_query_str}</query>
                        <pagesize>{page_size}</pagesize>
                        <returnFormat>xml</returnFormat>
                      </readByQuery>
                    </function>"""
            else:
                function_xml = f"""
                    <function controlid="page-{page_num}">
                      <readMore>
                        <resultId>{result_id}</resultId>
                      </readMore>
                    </function>"""

            payload = f"""<?xml version="1.0" encoding="UTF-8"?>
                <request>
                  <control>
                    <senderid>{SENDER_ID}</senderid>
                    <password>{SENDER_PASSWORD}</password>
                    <controlid>batch-page-{page_num}</controlid>
                    <uniqueid>false</uniqueid>
                    <dtdversion>3.0</dtdversion>
                  </control>
                  <operation>
                    <authentication>
                      <sessionid>{session_id}</sessionid>
                    </authentication>
                    <content>
                      {function_xml}
                    </content>
                  </operation>
                </request>"""

            response = await client.post(
                SAGE_URL, content=payload,
                headers={"Content-Type": "application/xml"}
            )

            control_check = xmltodict.parse(response.text)
            ctrl = control_check.get("response", {}).get("control", {})
            if ctrl.get("status") == "failure":
                err = control_check["response"]["errormessage"]["error"]["description2"]
                raise Exception(f"Sage control error [{object_name}]: {err}")

            data = xmltodict.parse(response.text, force_list=(record_tag,))
            result = data["response"]["operation"]["result"]

            if result["status"] == "failure":
                err = result["errormessage"]["error"]["description2"]
                raise Exception(f"Sage query error [{object_name}]: {err}")

            data_block = result["data"]

            if page_num == 1:
                debug_path = f"debug_page1_{object_name.lower()}.xml"
                with open(debug_path, "w") as f:
                    f.write(response.text)
                print(f"  ✓ Raw response saved to {debug_path}")

            if total_count is None:
                total_count = int(data_block.get("@totalcount", 0))
                print(f"  ℹ [{object_name}] Total records to fetch: {total_count}")

            num_remaining = int(data_block.get("@numremaining", 0))

            new_result_id = data_block.get("@resultId")
            if new_result_id:
                result_id = new_result_id

            records = data_block.get(record_tag, [])
            if isinstance(records, dict):
                records = [records]

            if not records and num_remaining > 0:
                raise Exception(
                    f"Parsed 0 {record_tag} records but numremaining={num_remaining}. "
                    f"data_block keys: {list(data_block.keys())}"
                )

            all_records.extend(records)

            print(f"  ✓ [{object_name}] Page {page_num}: +{len(records)} records "
                  f"| total: {len(all_records)}/{total_count} "
                  f"| remaining: {num_remaining}")

            if num_remaining == 0:
                print(f"  ✓ [{object_name}] Done — numremaining = 0")
                break
            if total_count > 0 and len(all_records) >= total_count:
                print(f"  ✓ [{object_name}] Done — fetched {len(all_records)}/{total_count}")
                break
            if len(records) == 0:
                print(f"  ✓ [{object_name}] Done — empty page returned")
                break
            if result_id is None:
                raise Exception(f"readByQuery [{object_name}] did not return a resultId — cannot paginate.")

            page_num += 1
            await asyncio.sleep(0.2)

    return all_records


# ── Query builders per object ───────────────────────────────────────────────

def build_query(object_name: str, financial_entity: str) -> str:
    """
    Field names confirmed from lookup_<OBJECT>.xml responses:
      - ARPYMT:   has FINANCIALENTITY. CLEARED valid values: T, F, M.
      - DEPOSIT:  has FINANCIALENTITY. CLEARED valid values: T, F, M.
      - OTHERRECEIPTS: NO FINANCIALENTITY field — uses BANKACCOUNTID instead.
      - FUNDSTRANSFER: NO single account field — uses FROMACCOUNTID / TOACCOUNTID
        since a transfer touches two accounts.

    CLEARED = 'F' filters to strictly uncleared rows (excludes 'T' cleared
    and 'M' matched-but-unconfirmed).

    ARPYMT also filters out BANKTRXAMOUNT = 0 — these are AR reapplication/
    credit-application entries (DOCNUMBER like "CREDIT APPLIED" or "OP OF
    ... APPLIED TO ...") that never actually moved money through the bank,
    which is why Sage's Bank Rec screen excludes them too.
    """
    if object_name == "FUNDSTRANSFER":
        #return (f"(FROMACCOUNTID = '{financial_entity}' "
        #        f"or TOACCOUNTID = '{financial_entity}') and CLEARED != 'T'")
        return (f"TOACCOUNTID = '{financial_entity}' and CLEARED != 'T'")
    elif object_name == "OTHERRECEIPTS":
        return f"BANKACCOUNTID = '{financial_entity}' and CLEARED = 'F'"
    elif object_name == "DEPOSIT":
        return f"FINANCIALENTITY = '{financial_entity}' and CLEARED = 'F'"
    elif object_name == "ARPYMT":
        return (f"FINANCIALENTITY = '{financial_entity}' and CLEARED != 'T' "
                f"and BANKTRXAMOUNT != 0")
    elif object_name == "APPYMT":
        # Same shape as ARPYMT, just the vendor-side (outbound) payment
        # object. BANKTRXAMOUNT != 0 excludes non-cash entries the same
        # way it does on ARPYMT (e.g. credit-application/reapplication
        # rows that never actually moved money through the bank).
        # CONFIRMED via sandbox query: FFB_4449 returned 103 uncleared
        # records with this filter, paginating correctly.
        return (f"FINANCIALENTITY = '{financial_entity}' and CLEARED != 'T' "
                f"and BANKTRXAMOUNT != 0")
    else:
        raise ValueError(f"Unknown object: {object_name}")


async def run_diagnostic_sample(session_id: str, object_name: str, account_field: str) -> None:
    """
    Pulls a handful of records with NO financial-entity filter, showing
    just RECORDNO / the account field / CLEARED, so you can see what
    real account-ID values and CLEARED states actually look like in this
    company before trusting a 0-result filtered query.
    """
    fields_str = f"RECORDNO,{account_field},CLEARED"
    payload = f"""<?xml version="1.0" encoding="UTF-8"?>
<request>
  <control>
    <senderid>{SENDER_ID}</senderid>
    <password>{SENDER_PASSWORD}</password>
    <controlid>diag-{object_name}</controlid>
    <uniqueid>false</uniqueid>
    <dtdversion>3.0</dtdversion>
  </control>
  <operation>
    <authentication>
      <sessionid>{session_id}</sessionid>
    </authentication>
    <content>
      <function controlid="diag-{object_name}">
        <readByQuery>
          <object>{object_name}</object>
          <fields>{fields_str}</fields>
          <query></query>
          <pagesize>5</pagesize>
          <returnFormat>xml</returnFormat>
        </readByQuery>
      </function>
    </content>
  </operation>
</request>"""

    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.post(
            SAGE_URL, content=payload,
            headers={"Content-Type": "application/xml"}
        )

    debug_path = f"diagnostic_{object_name.lower()}.xml"
    with open(debug_path, "w") as f:
        f.write(response.text)

    try:
        data = xmltodict.parse(response.text, force_list=(object_name.lower(),))
        result = data["response"]["operation"]["result"]
        if result.get("status") == "failure":
            err = result.get("errormessage", {}).get("error", {})
            print(f"  ✗ [{object_name}] diagnostic query failed: {err.get('description2', err)}")
            return
        data_block = result["data"]
        total = data_block.get("@totalcount", "?")
        records = data_block.get(object_name.lower(), [])
        if isinstance(records, dict):
            records = [records]
        print(f"  ℹ [{object_name}] {total} total records exist (no filter). "
              f"Sample {account_field}/CLEARED values:")
        for r in records:
            print(f"      RECORDNO={r.get('RECORDNO')}  "
                  f"{account_field}={r.get(account_field)!r}  "
                  f"CLEARED={r.get('CLEARED')!r}")
    except Exception as e:
        print(f"  ⚠ [{object_name}] could not parse diagnostic response: {e} "
              f"(see {debug_path})")


# ── Normalization ────────────────────────────────────────────────────────

def normalize_record(r: dict, source_object: str) -> dict:
    """
    Best-guess convenience keys per object — all UNVERIFIED until checked
    against the "raw" key. The full raw record is always preserved so
    nothing is silently dropped even if these guesses are wrong.
    """
    common = {
        "record_no":        r.get("RECORDNO"),
        "cleared":          r.get("CLEARED"),
        "clear_date":       r.get("CLRDATE"),
        "source_object":    source_object,
        "description":      None,   # overridden below where the object has one
        "party_name":       None,   # only ARPYMT has a customer/vendor party
        # DISPLAYSTATE flags things like "Reversed" (STATE='V') that a raw
        # BANKTRXAMOUNT filter alone won't catch — worth surfacing on every
        # object, not just ARPYMT, in case others carry the same field.
        # Both raw values kept separately — STATE is the short code (e.g.
        # "V"), DISPLAYSTATE is the human-readable label (e.g. "Reversed").
        "state":            r.get("STATE"),
        "display_state":    r.get("DISPLAYSTATE"),
        # Convenience flag (not a raw Intacct field) so the Excel export
        # can be sorted/filtered against the bank statement's debit vs
        # credit columns. Every source object here is a credit/inflow
        # except APPYMT, which overrides this below.
        "direction":        "inflow",
    }

    if source_object == "ARPYMT":
        common.update({
            "date":           r.get("RECEIPTDATE") or r.get("WHENPAID"),
            "doc_number":     r.get("DOCNUMBER"),
            "description":    r.get("DESCRIPTION"),
            "party_name":     r.get("CUSTOMERNAME") or r.get("VENDORNAME"),
            # BANKTRXAMOUNT is what Sage's Bank Rec screen shows as
            # "Bank amount"/"Txn amount" — this is real cash movement.
            # TOTALPAID/TOTALRECEIVED can be "0" (e.g. on advance/credit-card
            # receipts not yet applied to an invoice) even when real money
            # hit the bank, so they're not reliable here.
            "total":          r.get("BANKTRXAMOUNT"),
            "payment_method": r.get("PAYMENTMETHOD"),
            "financial_entity": r.get("FINANCIALENTITY"),
        })
    elif source_object == "OTHERRECEIPTS":
        common.update({
            "date":           r.get("WHENPAID") or r.get("DEPOSITDATE"),
            "doc_number":     r.get("DOCNUMBER"),
            "description":    r.get("DESCRIPTION") or r.get("DESCRIPTION2"),
            "total":          r.get("TOTALENTERED"),
            "bank_account":   r.get("BANKACCOUNTID"),
        })
    elif source_object == "DEPOSIT":
        common.update({
            "date":           r.get("BATCHDATE"),
            "doc_number":     r.get("DEPOSITID"),
            "description":    r.get("DESCRIPTION"),
            # Fallback to BANKTRXAMOUNT in case TOTALENTERED is ever 0 while
            # real cash moved — same class of issue found on ARPYMT.
            "total":          r.get("TOTALENTERED") or r.get("BANKTRXAMOUNT"),
            "financial_entity": r.get("FINANCIALENTITY"),
        })
    elif source_object == "FUNDSTRANSFER":
        common.update({
            "date":           r.get("WHENCREATED"),
            "doc_number":     r.get("RECORDID"),
            "description":    r.get("DESCRIPTION"),
            "total":          r.get("AMOUNT") or r.get("TOTALENTERED"),
            "from_account":   r.get("FROMACCOUNTID"),
            "to_account":     r.get("TOACCOUNTID"),
        })
    elif source_object == "APPYMT":
        common.update({
            "date":           r.get("PAYMENTDATE") or r.get("WHENPAID"),
            "doc_number":     r.get("DOCNUMBER"),
            "description":    r.get("DESCRIPTION"),
            "party_name":     r.get("VENDORNAME"),
            # BANKTRXAMOUNT mirrors ARPYMT's convention for "what actually
            # hit the bank." CONFIRMED: this value is POSITIVE for AP
            # payments even though it's an outflow (e.g. a $14,626.52
            # vendor check shows as BANKTRXAMOUNT=14626.52, not -14626.52).
            # Do not rely on the sign to detect debit vs. credit — use the
            # "direction" flag below instead when matching against the
            # bank statement.
            "total":          r.get("BANKTRXAMOUNT"),
            "payment_method": r.get("PAYMENTMETHOD"),
            "financial_entity": r.get("FINANCIALENTITY"),
            "direction":      "outflow",   # convenience flag, not a raw Intacct field
        })

    common["raw"] = r
    return common


# ── Excel export ─────────────────────────────────────────────────────────

def write_excel_report(transactions: list[dict], output_path: str) -> None:
    """
    Writes an .xlsx with one row per uncleared (CLEARED='F') transaction,
    columns: source_object, date, doc_number, description, total,
    party_name, payment_method, state, display_state, cleared.
    Raw data only — no formulas needed for a flat export like this.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Uncleared Transactions"

    headers = ["source_object", "direction", "date", "doc_number", "description",
               "total", "party_name", "payment_method", "state", "display_state",
               "cleared"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for txn in transactions:
        if txn.get("cleared") != "F":
            continue
        ws.append([
            txn.get("source_object"),
            txn.get("direction"),
            txn.get("date"),
            txn.get("doc_number"),
            txn.get("description"),
            txn.get("total"),
            txn.get("party_name"),
            txn.get("payment_method"),
            txn.get("state"),
            txn.get("display_state"),
            txn.get("cleared"),
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    # reasonable column widths
    widths = [16, 10, 14, 16, 40, 14, 30, 16, 10, 14, 10]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    wb.save(output_path)
    print(f"✓ Saved Excel report to: {output_path}")


# ── Main orchestration ──────────────────────────────────────────────────────

async def main():
    print("\n=== Sage Bank Reconciliation — All Uncleared Transactions Fetch ===\n")

    session_id = await get_session_id()

    print(f"\n--- Step 1: field discovery (lookup) for all {len(OBJECTS)} objects ---")
    for object_name in OBJECTS:
        await run_lookup(session_id, object_name)

    print(f"\n--- Step 2: diagnostic sample (no filter) for all {len(OBJECTS)} objects ---")
    print("    (confirms real account-ID values/formats before trusting a 0-result filter)")
    for object_name, account_field in ACCOUNT_FIELD.items():
        await run_diagnostic_sample(session_id, object_name, account_field)

    for financial_entity in FINANCIAL_ENTITIES:
        print(f"\n{'='*70}")
        print(f"Fetching for financial entity: {financial_entity}")
        print(f"{'='*70}")

        all_normalized = []
        counts = {}

        for object_name, record_tag in OBJECTS.items():
            query_str = build_query(object_name, financial_entity)
            print(f"\n[{object_name}] query: {query_str}")
            try:
                records = await fetch_object_records(
                    session_id=session_id,
                    object_name=object_name,
                    record_tag=record_tag,
                    query_str=query_str,
                )
            except Exception as e:
                print(f"  ✗ [{object_name}] fetch failed: {e}")
                print(f"    (check lookup_{object_name}.xml to confirm the right "
                      f"financial-entity field name and adjust build_query())")
                records = []

            counts[object_name] = len(records)
            all_normalized.extend(normalize_record(r, object_name) for r in records)

        print(f"\n{'-'*70}")
        print(f"[{financial_entity}] Totals: " +
              ", ".join(f"{k}={v}" for k, v in counts.items()) +
              f" | combined={len(all_normalized)}")
        print(f"{'-'*70}")

        output_path = f"uncleared_transactions_{financial_entity}.json"
        with open(output_path, "w") as f:
            json.dump({
                "financial_entity": financial_entity,
                "counts": counts,
                "total": len(all_normalized),
                "transactions": all_normalized,
            }, f, indent=2)
        print(f"✓ Saved results to: {output_path}")

        xlsx_path = f"uncleared_transactions_{financial_entity}.xlsx"
        write_excel_report(all_normalized, xlsx_path)


if __name__ == "__main__":
    asyncio.run(main())
